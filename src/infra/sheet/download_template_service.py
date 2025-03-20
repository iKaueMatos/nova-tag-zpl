# ----------------------------------------------------------------------------
# Autor: Kaue de Matos
# Empresa: Nova Software
# Propriedade da Empresa: Todos os direitos reservados
# ----------------------------------------------------------------------------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class TemplateDownloadService:
    def __init__(self, root):
        self.root = root

    def download_template(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Template de Dados"

        headers = ["EAN", "SKU", "Quantidade", "Código", "Descrição", "Tamanho"]
        example_data = [
            ["7891234567890", "SKU12345", 10, "ABCDEFDEE", "Produto Exemplo 1", "M"],
            ["7890987654321", "SKU54321", 20, "987654321", "Produto Exemplo 2", "G"],
            ["", "", "", "", "", ""]
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_num, row_data in enumerate(example_data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_data)

        instructions = (
            "Instruções:\n"
            "- Preencha a coluna 'EAN' com os códigos de barras dos produtos.\n"
            "- A coluna 'SKU' deve conter o identificador único do produto.\n"
            "- A coluna 'Quantidade' indica quantas etiquetas daquele EAN do produto serão geradas.\n"
            "- A coluna 'Código' deve conter um identificador exclusivo de 9 dígitos.\n"
            "- A coluna 'Descrição' pode conter detalhes adicionais do produto.\n"
            "- A coluna 'Tamanho' deve conter informações como P, M, G, etc.\n"
            "- Se não quiser preencher um dado, deixe em branco e o sistema irá substituir por '-'.\n"
        )
        sheet.cell(row=6, column=8, value=instructions)
        sheet.column_dimensions['H'].width = 80

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            try:
                workbook.save(file_path)
                messagebox.showinfo("Sucesso", f"Template salvo em: {file_path}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar o template: {e}")